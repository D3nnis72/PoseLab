#!/usr/bin/env python3
import json
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font

# --------------------------------------------------------------------------------
# 1) Your fixed folder lists
# --------------------------------------------------------------------------------
OBJECTS      = {"brake", "crankcase"}
VISIBILITIES = {"free"}
LIGHTINGS    = {"bright", "low"}
DISTANCES    = {"75cm", "150cm"}
HEIGHTS      = {"115h", "145h"}
ANGLES       = {"0", "60", "120", "180", "240", "300"}
METHODS      = {"FoundationPose", "MegaPose", "GigaPose", "OVE6D", "SAM6D"}

# --------------------------------------------------------------------------------
# 2) Collect all records under the exact folder hierarchy
# --------------------------------------------------------------------------------
def collect_records(root: Path):
    out = []
    for obj_dir in root.iterdir():
        if obj_dir.name not in OBJECTS or not obj_dir.is_dir(): continue
        for vis_dir in obj_dir.iterdir():
            if vis_dir.name not in VISIBILITIES or not vis_dir.is_dir(): continue
            for lig_dir in vis_dir.iterdir():
                if lig_dir.name not in LIGHTINGS or not lig_dir.is_dir(): continue
                for dist_dir in lig_dir.iterdir():
                    if dist_dir.name not in DISTANCES or not dist_dir.is_dir(): continue
                    for ht_dir in dist_dir.iterdir():
                        if ht_dir.name not in HEIGHTS or not ht_dir.is_dir(): continue
                        for ang_dir in ht_dir.iterdir():
                            if ang_dir.name not in ANGLES or not ang_dir.is_dir(): continue
                            for m_dir in ang_dir.iterdir():
                                if m_dir.name not in METHODS or not m_dir.is_dir(): continue
                                sf = m_dir / "Scores" / "scores_bop19.json"
                                tf = m_dir / "Scores" / "timings.json"
                                if sf.is_file() and tf.is_file():
                                    scores = json.load(open(sf))
                                    times  = json.load(open(tf))
                                    rec = dict(
                                        object=obj_dir.name,
                                        visibility=vis_dir.name,
                                        lighting=lig_dir.name,
                                        distance=dist_dir.name,
                                        height=ht_dir.name,
                                        angle=ang_dir.name,
                                        method=m_dir.name,
                                    )
                                    rec.update(scores)
                                    for k,v in times.items():
                                        rec[f"time_{k}"] = v
                                    out.append(rec)
    return out

# --------------------------------------------------------------------------------
# 3) Helpers
# --------------------------------------------------------------------------------
def write_pivot(ws, start_row, pivot, title):
    ws.cell(row=start_row, column=1, value=title).font = Font(bold=True)
    # headers
    for ci, method in enumerate(pivot.columns, start=2):
        ws.cell(row=start_row+1, column=ci, value=method).font = Font(bold=True)
    # data
    for ri, idx in enumerate(pivot.index, start=start_row+2):
        ws.cell(row=ri, column=1, value=idx)
        for ci, method in enumerate(pivot.columns, start=2):
            ws.cell(row=ri, column=ci, value=float(pivot.at[idx, method]))
    return start_row + 1 + len(pivot)

def add_chart(ws, table_row, table_rows, table_cols, pos, title, offset):
    chart = BarChart()
    chart.title = title
    chart.y_axis.title = "Value"
    # include header row + all data rows
    data = Reference(
        ws,
        min_col=2,
        max_col=1+table_cols,
        min_row=table_row+1,
        max_row=table_row+offset+table_rows
    )
    # categories are just the data labels (skip header row)
    cats = Reference(
        ws,
        min_col=1,
        min_row=table_row+2,
        max_row=table_row+offset+table_rows
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, pos)


# --------------------------------------------------------------------------------
# 4) Build detailed report
# --------------------------------------------------------------------------------
def write_detailed_report(df: pd.DataFrame, out: Path):
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    dims    = ["object","visibility","lighting","distance","height","angle"]
    score_metrics = sorted(c for c in df.columns if c.startswith("bop19_"))
    time_metrics  = sorted(c for c in df.columns if c.startswith("time_"))
    metrics = score_metrics + time_metrics

    for dim in dims:
        for m in metrics:
            pivot = df.pivot_table(index=dim, columns="method", values=m, aggfunc="mean")\
                      .fillna(0).round(3)
            sheet = wb.create_sheet(f"{dim}_{m}")
            # overall
            row0 = write_pivot(sheet, 1, pivot, f"{m} by {dim} (overall)")
            # if not object, add breakouts per object
            if dim != "object":
                for obj in ["brake","crankcase"]:
                    pivot_obj = df[df.object==obj]\
                                  .pivot_table(index=dim, columns="method", values=m, aggfunc="mean")\
                                  .fillna(0).round(3)
                    row0 += 2
                    row0 = write_pivot(sheet, row0, pivot_obj, f"{m} by {dim} ({obj})")
                total_rows = row0 - 1
                # charts: overall at K2, brake at K20, crankcase at K38
                add_chart(sheet, 1, len(pivot), len(pivot.columns), "K2",  f"{m} overall", 1)
                offset = -1
                if dim == "visibility":
                    print("Setting visibility offset to -1")
                    offset = -2
                if dim == 'angle':
                    offset += 3

                add_chart(sheet, len(pivot)+4, len(pivot.columns), len(pivot.columns), "K20", f"{m} brake", offset)
                add_chart(sheet, len(pivot)*2+7, len(pivot.columns), len(pivot.columns), "K38", f"{m} crankcase", offset)
            else:
                # single chart for object dimension
                add_chart(sheet, 1, len(pivot), len(pivot.columns), "K2", f"{m} by object", 1)

    wb.save(out)
    print(f"✅ Detailed report written to {out}")

# --------------------------------------------------------------------------------
# 5) CLI
# --------------------------------------------------------------------------------
def main():
    import argparse
    parser = argparse.ArgumentParser(
        description="Aggregate JSON scores+timings into Excel with rich charts"
    )
    parser.add_argument(
        "--result_root", type=Path, default=Path("./Result"),
        help="Root folder containing brake/crankcase subdirs"
    )
    parser.add_argument(
        "--output", type=Path, default=Path("detailed_report.xlsx"),
        help="Output Excel file path"
    )
    args = parser.parse_args()

    if not args.result_root.is_dir():
        print(f"ERROR: {args.result_root} is not a directory")
        return

    records = collect_records(args.result_root)
    if not records:
        print(f"No JSON records found under {args.result_root}")
        return

    df = pd.DataFrame(records)
    base_cols = ["object","visibility","lighting","distance","height","angle","method"]
    score_cols = sorted(c for c in df.columns if c.startswith("bop19_"))
    time_cols  = sorted(c for c in df.columns if c.startswith("time_"))
    df = df[base_cols + score_cols + time_cols]

    write_detailed_report(df, args.output)

if __name__=="__main__":
    main()
