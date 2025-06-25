#!/usr/bin/env python3
import json
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font

# fixed folders
OBJECTS      = {"brake", "crankcase"}
VISIBILITIES = {"free"}
LIGHTINGS    = {"bright", "low"}
DISTANCES    = {"75cm", "150cm"}
HEIGHTS      = {"115h", "145h"}
ANGLES       = {"0", "60", "120", "180", "240", "300"}
METHODS      = {"FoundationPose", "MegaPose", "GigaPose", "OVE6D", "SAM6D"}

def collect_records(root: Path):
    out = []
    for obj_dir in root.iterdir():
        if obj_dir.name not in OBJECTS or not obj_dir.is_dir(): continue
        for vis_dir in obj_dir.iterdir():
            if vis_dir.name not in VISIBILITIES or not vis_dir.is_dir(): continue
            for lig_dir in vis_dir.iterdir():
                if lig_dir.name not in LIGHTINGS or not lig_dir.is_dir(): continue
                for dist_dir in lig_dir.iterdir():
                    if dist_dir.name not in DISTANCES or not dist_dir.is_dir(): continue
                    for ht_dir in dist_dir.iterdir():
                        if ht_dir.name not in HEIGHTS or not ht_dir.is_dir(): continue
                        for ang_dir in ht_dir.iterdir():
                            if ang_dir.name not in ANGLES or not ang_dir.is_dir(): continue
                            for m_dir in ang_dir.iterdir():
                                if m_dir.name not in METHODS or not m_dir.is_dir(): continue
                                sf = m_dir/"Scores"/"scores_bop19.json"
                                tf = m_dir/"Scores"/"timings.json"
                                if sf.is_file() and tf.is_file():
                                    scores = json.load(open(sf))
                                    timings= json.load(open(tf))
                                    rec = {
                                        "object": obj_dir.name,
                                        "visibility": vis_dir.name,
                                        "lighting": lig_dir.name,
                                        "distance": dist_dir.name,
                                        "height": ht_dir.name,
                                        "angle": ang_dir.name,
                                        "method": m_dir.name,
                                    }
                                    rec.update(scores)
                                    for k,v in timings.items():
                                        rec[f"time_{k}"] = v
                                    out.append(rec)
    return out

def add_bar_chart(ws, df, index_col, value_col, title, cell):
    chart = BarChart()
    chart.title = title
    chart.y_axis.title = value_col
    cats = Reference(ws, min_col=1, min_row=2, max_row=1+df.shape[0])
    data = Reference(ws, min_col=2, min_row=1, max_row=1+df.shape[0])
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, cell)

def write_excel(df, output: Path):
    wb = Workbook()
    # ---- All Results ----
    ws = wb.active; ws.title="All Results"
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    for c in ws["1:1"]: c.font=Font(bold=True)

    # ---- Avg Times ----
    df_time = df.groupby("method")["time_total"].mean().reset_index()
    ws2 = wb.create_sheet("Avg Times")
    for r in dataframe_to_rows(df_time, index=False, header=True): ws2.append(r)
    for c in ws2["1:1"]: c.font=Font(bold=True)
    add_bar_chart(ws2, df_time, "method", "time_total", "Avg Total Time", "D2")

    # ---- Performance sheets ----
    dims = ["object","visibility","lighting","height","angle"]
    for dim in dims:
        df_perf = df.groupby([dim,"method"])["bop19_average_recall"]\
                    .mean().unstack().fillna(0)
        wsn = wb.create_sheet(f"Perf by {dim.title()}")
        # write header
        hdr = [dim.title()] + list(df_perf.columns)
        wsn.append(hdr)
        for idx, row in df_perf.iterrows():
            wsn.append([idx] + row.tolist())
        for c in wsn["1:1"]: c.font=Font(bold=True)
        # chart
        add_bar_chart(wsn, df_perf, dim, "bop19_average_recall",
                      f"Avg Recall by Method per {dim.title()}", "H2")

    wb.save(output)
    print(f"Report saved to {output}")

def main():
    import argparse
    p = argparse.ArgumentParser()
    p.add_argument("--result_root", type=Path, default=Path("./Result"))
    p.add_argument("--output",      type=Path, default=Path("report.xlsx"))
    args = p.parse_args()

    if not args.result_root.is_dir():
        print("Invalid root:", args.result_root); return

    recs = collect_records(args.result_root)
    if not recs:
        print("No data found."); return

    df = pd.DataFrame(recs)
    cols = ["object","visibility","lighting","distance","height","angle","method"]
    score_cols = [c for c in df if c.startswith("bop19_")]
    time_cols  = [c for c in df if c.startswith("time_")]
    df = df[cols + score_cols + time_cols]

    write_excel(df, args.output)

if __name__=="__main__":
    main()